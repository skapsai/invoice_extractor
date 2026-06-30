"""
invoice_extractor.py
--------------------
Extracts invoice data from PDF files (text-based or scanned/OCR)
and writes results to a formatted Excel file.

Supports any vendor using a strict generic extraction approach:
  - Table-based extraction (preferred)
  - Pattern-based fallback (if no table found)
  - Multi-invoice PDF support (splits by page / invoice number)
  - Noise filtering to prevent table borders, tax lines, bank details,
    headers, and footers from becoming item rows.

Folder structure:
    project_folder/
        invoice_extractor.py
        invoices/             <- Place PDF invoices here
        extracted_text/       <- Raw extracted text saved here (auto-created)
        output/               <- Excel output saved here (auto-created)
        processed/            <- Successfully processed PDFs moved here
        error/                <- Failed PDFs moved here

Requirements:
    pip install pdfplumber pytesseract pdf2image pillow pandas openpyxl

External dependencies:
    - Tesseract OCR : https://github.com/UB-Mannheim/tesseract/wiki
    - Poppler       : https://github.com/oschwartz10612/poppler-windows/releases
                      Add Poppler bin to PATH, or set POPPLER_PATH below.
"""

import os
import json
import re
import time
import sys
import shutil
import logging
from pathlib import Path
from typing import Optional

import requests
import pdfplumber
import pytesseract
from pdf2image import convert_from_path
from PIL import Image  # noqa: F401 – required by pdf2image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION & PORTABILITY (Supports PyInstaller .exe builds)
# ─────────────────────────────────────────────────────────────────────────────

# Determine base directory dynamically based on whether running as compiled .exe or script
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).parent

def _load_env_configs() -> dict:
    configs = {
        "PADDLE_OCR_TOKEN": "e626136dd8ef9f6ca72a4515f346783d1d8fc003",
        "TESSERACT_CMD": "",
        "POPPLER_PATH": ""
    }
    
    # Try importing dotenv first, else fallback to reading manually
    try:
        from dotenv import load_dotenv
        load_dotenv(BASE_DIR / ".env")
    except ImportError:
        pass
        
    dotenv_path = BASE_DIR / ".env"
    if dotenv_path.exists():
        try:
            with open(dotenv_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    if "=" in line:
                        k, v = line.split("=", 1)
                        key = k.strip()
                        val = v.strip().strip('"').strip("'")
                        configs[key] = val
        except Exception:
            pass

    # Override with system environment variables if they exist
    for key in configs:
        if key in os.environ:
            configs[key] = os.environ[key]
            
    return configs

_env_configs = _load_env_configs()

# Tesseract executable configuration
_tess_env = _env_configs.get("TESSERACT_CMD", "")
_tess_local = BASE_DIR / "Tesseract-OCR" / "tesseract.exe"
_tess_std = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")
_tess_dev = Path(r"D:\Softwares\Tesseract-ocr\tesseract.exe")

if _tess_env:
    pytesseract.pytesseract.tesseract_cmd = _tess_env
elif _tess_local.exists():
    pytesseract.pytesseract.tesseract_cmd = str(_tess_local)
elif _tess_std.exists():
    pytesseract.pytesseract.tesseract_cmd = str(_tess_std)
elif _tess_dev.exists():
    pytesseract.pytesseract.tesseract_cmd = str(_tess_dev)
else:
    # Let it default to system PATH
    pass

# Poppler path configuration
_pop_env = _env_configs.get("POPPLER_PATH", "")
_pop_local = BASE_DIR / "poppler" / "bin"
_pop_dev = r"D:\Softwares\Poppler\poppler-26.02.0\Library\bin"

if _pop_env:
    POPPLER_PATH = _pop_env
elif _pop_local.exists():
    POPPLER_PATH = str(_pop_local)
elif Path(_pop_dev).exists():
    POPPLER_PATH = _pop_dev
else:
    POPPLER_PATH = None

# Paddle-OCR API settings
USE_PADDLE_OCR = True  # Set to True to use Paddle-OCR API as primary OCR fallback
PADDLE_OCR_TOKEN = _env_configs.get("PADDLE_OCR_TOKEN", "e626136dd8ef9f6ca72a4515f346783d1d8fc003")
PADDLE_OCR_MODEL = "PaddleOCR-VL-1.6"
PADDLE_OCR_JOB_URL = "https://paddleocr.aistudio-app.com/api/v2/ocr/jobs"

OCR_DPI = 300
MIN_TEXT_LENGTH = 100   # chars; below this → treat PDF as scanned


INVOICES_DIR       = BASE_DIR / "invoices"
EXTRACTED_TEXT_DIR = BASE_DIR / "extracted_text"
OUTPUT_DIR         = BASE_DIR / "output"
PROCESSED_DIR      = BASE_DIR / "processed"
ERROR_DIR          = BASE_DIR / "error"
OUTPUT_EXCEL       = OUTPUT_DIR / "extracted_invoice_records.xlsx"
OUTPUT_TXT         = OUTPUT_DIR / "extracted_invoice_records.txt"
LOG_FILE           = BASE_DIR / "invoice_extraction.log"

# ─────────────────────────────────────────────────────────────────────────────
# LOGGING  (console + append to log file)
# ─────────────────────────────────────────────────────────────────────────────

_fmt    = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s",
                             datefmt="%Y-%m-%d %H:%M:%S")
_file_h = logging.FileHandler(LOG_FILE, encoding="utf-8", mode="a")
_file_h.setFormatter(_fmt)
_con_h  = logging.StreamHandler()
_con_h.setFormatter(_fmt)

logger = logging.getLogger("invoice_extractor")
logger.setLevel(logging.DEBUG)
logger.addHandler(_file_h)
logger.addHandler(_con_h)

# Prevent verbose third-party libraries from polluting the logs
logging.getLogger("pdfminer").setLevel(logging.WARNING)
logging.getLogger("pdfplumber").setLevel(logging.WARNING)
logging.getLogger("PIL").setLevel(logging.WARNING)


def handle_exception(exc_type, exc_value, exc_traceback):
    """Log unhandled exceptions to the log file before exit."""
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    logger.critical("Unhandled exception occurred",
                    exc_info=(exc_type, exc_value, exc_traceback))


sys.excepthook = handle_exception


def log_and_print(msg: str) -> None:
    """Print to console and append directly to log file (for summary block)."""
    print(msg)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(msg + "\n")
    except Exception as e:
        logger.error(f"Failed to write to log file: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL COLUMNS
# ─────────────────────────────────────────────────────────────────────────────

EXCEL_COLUMNS = [
    "Source File",
    "Vendor Name",          # NEW: extracted from invoice header
    "Invoice Date",
    "Invoice Number",
    "PO Number",
    "Item",
    "HSN Number",           # renamed from Serial Number; now extracts HSN
    "Quantity",
    "Basic Price excluding GST",
    "Extraction Status",
]

# ─────────────────────────────────────────────────────────────────────────────
# COMPILED REGEX CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

# Page separator inserted by extract_text_ocr()
PAGE_SEP = re.compile(r"---\s*Page\s+\d+\s*---", re.IGNORECASE)

# ── Invoice header patterns ────────────────────────────────────────────────

_INV_DATE_PAT = re.compile(
    r"(?:Invoice\s*Date|Inv\.?\s*Date|Bill\s*Date|Document\s*Date|Dated)"
    r"\s*[:\-]?\s*"
    r"(\d{1,2}[\-/\.]\d{1,2}[\-/\.]\d{2,4}"
    r"|\d{1,2}[\-/\.][A-Za-z]+[\-/\.]\d{2,4}"
    r"|\d{1,2}[\-/\.\s][A-Za-z]+[\-/\.\s]\d{2,4}"
    r"|\d{1,2}\s+[A-Za-z]+\s+\d{4}"
    r"|\d{1,2}-[A-Za-z]+-\d{2,4})",
    re.IGNORECASE,
)

# Invoice number: allow spaces around slashes (e.g. "AMAR2512 / 216")
_INV_NUM_PAT = re.compile(
    r"(?:Invoice\s*(?:No\.?|Number)|Inv\.?\s*No\.?|Bill\s*No\.?|"
    r"Tax\s*Invoice\s*No\.?|Invoice\s*No\s*:)"
    r"\s*[:\-]?\s*(?:[/|\\\-]\s*)?([A-Z0-9][A-Z0-9/_\- \t]*[A-Z0-9]|[A-Z0-9]+)",
    re.IGNORECASE,
)
# "invoiceNo.—— HLE/26-27/87" style
_INV_NUM_PAT2 = re.compile(
    r"invoiceNo\.?\s*[-–—]+\s*([A-Z0-9][A-Z0-9/_\-]*)",
    re.IGNORECASE,
)
# Bare invoice number on its own line matching known patterns like "ATPL/26-27/00417"
_INV_NUM_BARE = re.compile(
    r"^\s*([A-Z]{2,}/\d{2}-\d{2}/\d{3,}|[A-Z]{2,}-\d{2}-\d{2}-\d{3,})\s*$",
    re.IGNORECASE,
)

# PO Number: many label aliases, value on same line OR next line
_PO_LABELS = re.compile(
    r"(?:P\.?O\.?\s*Ref\.?\s*No\.?|P\.?O\.?\s*No\.?|PO\s*Number|Purchase\s*Order\s*(?:No\.?)?|"
    r"Buyer'?s?\s*Order\s*No\.?|Customer\s*Ref(?:erence)?\s*(?:No\.?)?|"
    r"Order\s*No\.?|P\.O\.\s*No\.?|P\.O\.No\.?)\s*[:\-]?\s*",
    re.IGNORECASE,
)
_PO_VALUE = re.compile(r"([A-Z0-9][A-Z0-9/_\-]{3,})", re.IGNORECASE)

# Vendor name: first prominent company/enterprise name in the document
_VENDOR_KEYWORDS = re.compile(
    r"((?:[A-Z][A-Za-z0-9&.'\-]+[ \t]+){1,6}"
    r"(?:Pvt\.?\s*Ltd\.?|Ltd\.?|Limited|Technologies|Enterprises?|"
    r"Solutions?|Industries|Barcode|Traders?|Services?|Corp\.?|"
    r"International|Supplies|Distributors?))",
    re.IGNORECASE,
)

# ── Table header detection ─────────────────────────────────────────────────

_TH_ITEM   = re.compile(
    r"description|particulars|name\s*of\s*(?:product|service|goods)|item",
    re.IGNORECASE,
)
_TH_QTY    = re.compile(
    r"\bqty\b|\bquantity\b|\bnos\b|\bpcs\b|\bunits?\b", re.IGNORECASE
)
_TH_AMOUNT = re.compile(
    r"\bamount\b|\brate\b|\bprice\b|\btaxable", re.IGNORECASE
)

# ── Table footer – stop reading rows at these ──────────────────────────────

_FOOTER_PAT = re.compile(
    r"^\s*(?:total|grand\s*total|sub\s*total|amount\s*in\s*words|"
    r"amount\s*chargeable|total\s*before\s*tax|taxable\s*amount|"
    r"tax\s*amount|total\s*amount|e\s*\.\s*&\s*o|add\s*:\s*[cs]gst|"
    r"sub\s*total\s*:?|remarks\s*:?)",
    re.IGNORECASE,
)

# ── Monetary value ─────────────────────────────────────────────────────────
# Matches numbers like 1,750.00  or  65,000.00  or  1366695.00
_MONEY_PAT = re.compile(r"\b[\d,]+\.\d{1,2}\b")

_STANDALONE_NUM_PAT = re.compile(
    r"(?<![\w\-\/])\d{1,3}(?:,\d{3})*(?:\.\d+)?(?![\w])|(?<![\w\-\/])\d+(?:\.\d+)?(?![\w])"
)

# ── HSN / SAC code: 4–8 digit standalone number ────────────────────────────
# Widened to 4 digits to catch codes like "9973" (service HSN)
_HSN_PAT = re.compile(r"\b(\d{4,8})\b")

# Alphanumeric HSN codes (some vendors use like "997331", "998434")
_HSN_ALPHA_PAT = re.compile(
    r"(?:HSN\s*(?:Code)?\s*[:\-]?|SAC\s*(?:Code)?\s*[:\-]?)\s*([A-Z0-9]{4,10})",
    re.IGNORECASE,
)

# ── Units of measure ──────────────────────────────────────────────────────
_UNIT_PAT = re.compile(
    r"\b(NOS|PCS|UNIT|UNITS|EA|SET|KG|MTR|LTR|BOX|PC|ROLL|EACH|NO\.|"
    r"MONTH|MONTHS|YEAR|YEARS|LICENSE|LICENSES|NAS)\b",
    re.IGNORECASE,
)

# ── Noise phrases ──────────────────────────────────────────────────────────
_NOISE_PHRASES = re.compile(
    r"tax\s*amount|amount\s*in\s*words?|bank\s*(?:details?|name|a/?c)|"
    r"ifsc|ifs\s*code|branch\s*&|authoris[e]?d\s*signatory|"
    r"authorized\s*signatory|declaration|terms\s*(and|&)\s*conditions?|"
    r"customer\s*signature|gstin\s*[:/]|pan\s*no|pan\s*:|round\s*off|"
    r"grand\s*total|total\s*amount\s*after\s*tax|total\s*before\s*tax|"
    r"total\s*taxable|add\s*:\s*[cs]gst|output\s*igst|hsn\s*/\s*sac\b|"
    r"taxable\s*value\s*cgst|rate\s*amount\s*tax|"
    r"page\s*\d+\s*of\s*\d+|e\.?&\.?o\.?e|"
    r"tax\s*invoice|original\s*for\s*recipient|computer\s*generated|"
    r"certified\s*that|goods\s*once\s*sold|interest\s*@|jurisdiction|"
    r"amount\s*chargeable|ack\s*(no|date)\s*:|irn\s*[-–]|upi\s*id|"
    r"cgst\s*:|sgst\s*:|igst\s*@|total\s*tax\b|dispatch\s*through|"
    r"payment\s*terms|delivery\s*at|contact\s*(person|no)|"
    r"place\s*of\s*supply|acc(?:ount)?\s*number|due\s*date|"
    r"bill\s*to\s*:|ship\s*to\s*:|regd\.?\s*address|"
    r"sales\s*by\s*:|income\s*tax\s*declaration|tds\s*on|"
    r"company'?s?\s*bank|e-invoice\s*irn|ack\s*no\.|ack\s*date|"
    r"we\s*declare|this\s*is\s*a\s*computer",
    re.IGNORECASE,
)

# Table-header row detector
_THEAD_ROW = re.compile(
    r"(?:sr\.?\s*no|sl\.?\s*no|s\.?\s*no|sno|si\.?\s*no)[\.\s]+"
    r"(?:description|particulars|name|item)",
    re.IGNORECASE,
)

_DESC_TRUNCATE_AT = re.compile(
    r"\b(?:Contract\s+Period|Contrat\s+Period|Model\s+No\.?|Machine\s+Serial|Serial\s+No\.?|"
    r"Last\s+Reading|Meter\s+Reading|Opening\s+Reading|Closing\s+Reading)\b",
    re.IGNORECASE,
)


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1: NOISE FILTER
# ─────────────────────────────────────────────────────────────────────────────

def looks_like_noise(text: str) -> bool:
    """
    Return True if text is noise that should NOT become an item row.
    Catches: table borders, OCR garbage, tax lines, totals, bank details,
    signatures, page footers, GST headers, and short/empty strings.
    """
    t = text.strip()
    if not t:
        return True
    # Filter amount in words (e.g. Rupees Only)
    if re.search(r"\b(rupees|only)\b", t, re.IGNORECASE):
        return True
    alnum = re.sub(r"[^a-zA-Z0-9]", "", t)
    if len(alnum) < 3:
        return True                          # nearly no real content
    if len(alnum) / max(len(t), 1) < 0.25:
        return True                          # mostly punctuation / box-drawing
    if _NOISE_PHRASES.search(t):
        return True
    if _THEAD_ROW.search(t):
        return True
    return False


# ─────────────────────────────────────────────────────────────────────────────
# STEP 2: TEXT EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def clean_text(text: str) -> str:
    """Normalize whitespace and remove common OCR artefacts."""
    text = re.sub(r"\r\n", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def extract_text_pdfplumber(pdf_path: Path) -> str:
    """Extract text from a text-based PDF using pdfplumber (page-by-page with markers)."""
    parts = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for i, page in enumerate(pdf.pages):
            t = page.extract_text()
            if t:
                parts.append(f"--- Page {i + 1} ---\n{t}")
    return "\n".join(parts)


def extract_text_ocr(pdf_path: Path, dpi: int = OCR_DPI) -> str:
    """Convert each PDF page to an image and run Tesseract OCR."""
    logger.info(f"  OCR: converting pages at {dpi} DPI …")
    kwargs: dict = {"dpi": dpi}
    if POPPLER_PATH:
        kwargs["poppler_path"] = POPPLER_PATH
    images = convert_from_path(str(pdf_path), **kwargs)
    parts = []
    for i, img in enumerate(images):
        parts.append(f"--- Page {i + 1} ---\n{pytesseract.image_to_string(img, lang='eng')}")
    return "\n".join(parts)


def extract_text_paddleocr(pdf_path: Path) -> str:
    """Submit PDF/image to Paddle-OCR API and get layout text with page separators."""
    logger.info(f"  Paddle-OCR: submitting '{pdf_path.name}' to API ...")
    
    headers = {
        "Authorization": f"bearer {PADDLE_OCR_TOKEN}",
    }
    
    optional_payload = {
        "useDocOrientationClassify": False,
        "useDocUnwarping": False,
        "useChartRecognition": False,
    }

    if not pdf_path.exists():
        raise FileNotFoundError(f"Local file not found: {pdf_path}")

    data = {
        "model": PADDLE_OCR_MODEL,
        "optionalPayload": json.dumps(optional_payload)
    }

    with open(pdf_path, "rb") as f:
        files = {"file": f}
        response = requests.post(PADDLE_OCR_JOB_URL, headers=headers, data=data, files=files)

    if response.status_code != 200:
        raise requests.HTTPError(f"Submission failed (HTTP {response.status_code}): {response.text}")

    resp_json = response.json()
    if resp_json.get("code") != 0 or "data" not in resp_json:
        raise ValueError(f"API returned error status ({resp_json.get('code')}): {resp_json.get('msg', 'Unknown error')}")

    job_id = resp_json["data"]["jobId"]
    logger.info(f"  Paddle-OCR: Job ID: {job_id}. Polling status...")
    
    start_time = time.time()
    
    while True:
        status_resp = requests.get(f"{PADDLE_OCR_JOB_URL}/{job_id}", headers=headers)
        if status_resp.status_code != 200:
            logger.warning(f"  Paddle-OCR: Failed to fetch job status (HTTP {status_resp.status_code}). Retrying in 5s...")
            time.sleep(5)
            continue
            
        status_json = status_resp.json()
        if status_json.get("code") != 0 or "data" not in status_json:
            raise ValueError(f"Status check failed: {status_json.get('msg', 'Unknown error')}")
            
        job_data = status_json["data"]
        state = job_data.get("state")
        elapsed = int(time.time() - start_time)
        
        if state == 'pending':
            logger.debug(f"  Paddle-OCR: Job pending... ({elapsed}s)")
        elif state == 'running':
            progress = job_data.get("extractProgress", {})
            total = progress.get("totalPages", "?")
            extracted = progress.get("extractedPages", "0")
            logger.debug(f"  Paddle-OCR: running, progress: {extracted}/{total} pages ({elapsed}s)")
        elif state == 'done':
            progress = job_data.get("extractProgress", {})
            extracted = progress.get("extractedPages", "0")
            logger.info(f"  Paddle-OCR: Job completed. {extracted} page(s) extracted.")
            jsonl_url = job_data.get("resultUrl", {}).get("jsonUrl")
            if not jsonl_url:
                raise ValueError("JSON URL not found in completed job response.")
            break
        elif state == 'failed':
            error_msg = job_data.get("errorMsg", "Unknown failure reason")
            raise RuntimeError(f"Paddle-OCR job failed: {error_msg}")
        
        time.sleep(5)

    logger.debug(f"  Paddle-OCR: Downloading result from Bos...")
    jsonl_resp = requests.get(jsonl_url)
    jsonl_resp.raise_for_status()
    
    lines = jsonl_resp.text.strip().split('\n')
    extracted_pages = []
    
    for line_num, line in enumerate(lines, start=1):
        line = line.strip()
        if not line:
            continue
        try:
            line_data = json.loads(line)
            result = line_data.get("result", {})
            layout_results = result.get("layoutParsingResults", [])
            for res in layout_results:
                page_text = res.get("markdown", {}).get("text", "")
                extracted_pages.append(page_text)
        except Exception as e:
            logger.error(f"  Paddle-OCR: Failed to parse result page line {line_num}: {e}")

    # Format with standard page separators
    formatted_text = ""
    for idx, page_text in enumerate(extracted_pages, start=1):
        formatted_text += f"--- Page {idx} ---\n{page_text}\n\n"
        
    return formatted_text.strip()


def split_spanned_text(text: str, rowspan: int) -> list[str]:
    """Splits a multi-line spanned cell text into `rowspan` separate item descriptions."""
    # Split text into lines, keeping track of original HTML line breaks
    # We clean up nested tags first or split by \n
    text_clean = text.replace("\\n", "\n").replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
    lines = [line.strip() for line in text_clean.split('\n') if line.strip()]
    if rowspan <= 1 or len(lines) <= 1:
        return [text] * rowspan
        
    sections = []
    current_section = []

    # Detail lines follow an item name and should remain attached to it.
    # A line is considered a "detail" if it matches reading info, dates, serial/model,
    # contract periods, or generic sub-descriptions like "One Month Rental".
    is_detail_pat = re.compile(
        r"\b(?:Last\s+Reading|Current\s+Reading|Opening\s+Reading|Closing\s+Reading|"
        r"Date\s*[;:]|Model\s+No|Machine\s+Serial|Serial\s+No|Contr?a?ct\s+Period)\b",
        re.IGNORECASE
    )

    # Lines that clearly start a new item (all-caps short words, or known billing keywords)
    # but are NOT detail lines should trigger a section break.
    for line in lines:
        line_clean = re.sub(r"<[^>]+>", "", line).strip()
        if not line_clean:
            continue

        is_detail = is_detail_pat.search(line_clean) is not None

        if current_section and not is_detail:
            # Start a new section only if we haven't reached rowspan limit
            if len(sections) + 1 < rowspan:
                sections.append(" ".join(current_section))
                current_section = []

        current_section.append(line_clean)

    if current_section:
        sections.append(" ".join(current_section))

    # Pad sections to match rowspan
    while len(sections) < rowspan:
        sections.append(sections[-1] if sections else "")

    return sections[:rowspan]


def convert_html_tables_to_text(text: str) -> str:
    """
    Finds HTML table elements in the text and converts them to pipe-separated plain text lines.
    This preserves the tabular layout with rowspan/colspan support.
    """
    table_pattern = re.compile(r"<table\b[^>]*>(.*?)</table>", re.DOTALL | re.IGNORECASE)
    row_pattern = re.compile(r"<tr\b[^>]*>(.*?)</tr>", re.DOTALL | re.IGNORECASE)
    cell_pattern = re.compile(r"<(t[dh])\b([^>]*?)>(.*?)</\1>", re.DOTALL | re.IGNORECASE)
    html_tag_pattern = re.compile(r"<[^>]+>")

    def replace_table(match):
        table_content = match.group(1)
        rows = row_pattern.findall(table_content)
        if not rows:
            return ""
            
        # Grid representation: row_idx -> col_idx -> text
        grid = {}
        
        for row_idx, row in enumerate(rows):
            if row_idx not in grid:
                grid[row_idx] = {}
                
            cells = cell_pattern.findall(row)
            
            col_idx = 0
            for tag, attrs, cell_val in cells:
                # Find the next free column in the current row
                while col_idx in grid[row_idx]:
                    col_idx += 1
                    
                # Parse rowspan / colspan
                rowspan_match = re.search(r'\browspan\s*=\s*["\']?(\d+)["\']?', attrs, re.IGNORECASE)
                rowspan = int(rowspan_match.group(1)) if rowspan_match else 1
                
                colspan_match = re.search(r'\bcolspan\s*=\s*["\']?(\d+)["\']?', attrs, re.IGNORECASE)
                colspan = int(colspan_match.group(1)) if colspan_match else 1
                
                # Split cell text across rowspan sections
                spanned_texts = split_spanned_text(cell_val, rowspan)
                
                # Fill grid for rowspan and colspan
                for r in range(rowspan):
                    target_row = row_idx + r
                    if target_row not in grid:
                        grid[target_row] = {}
                    
                    section_val = spanned_texts[r]
                    cell_text = html_tag_pattern.sub(" ", section_val)
                    cell_text = cell_text.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">").replace("&quot;", '"')
                    cell_text = re.sub(r"\s+", " ", cell_text).strip()
                    
                    for c in range(colspan):
                        grid[target_row][col_idx + c] = cell_text
                        
                col_idx += colspan

        # Build plain text rows
        plain_rows = []
        for r_idx in sorted(grid.keys()):
            row_data = grid[r_idx]
            if not row_data:
                continue
            cols = [row_data[c] for c in sorted(row_data.keys())]
            plain_rows.append(" | ".join(cols))
            
        return "\n".join(plain_rows)

    return table_pattern.sub(replace_table, text)


def get_pdf_text(pdf_path: Path) -> tuple[str, str, str]:
    """
    Try pdfplumber first; fall back to OCR if text is too short.
    If USE_PADDLE_OCR is True, uses Paddle-OCR API with local Tesseract fallback.
    Returns (text_300, text_200, method) where method ∈ {'pdfplumber', 'ocr', 'paddleocr'}.
    """
    text = extract_text_pdfplumber(pdf_path)
    # Strip page markers when counting useful content
    bare = PAGE_SEP.sub("", text).strip()
    if len(bare) >= MIN_TEXT_LENGTH:
        logger.info(f"  Extracted via pdfplumber ({len(bare)} chars).")
        return text, text, "pdfplumber"
        
    if USE_PADDLE_OCR:
        logger.info(f"  pdfplumber text too short ({len(bare)} chars) → Paddle-OCR API.")
        try:
            ocr_text = extract_text_paddleocr(pdf_path)
            return ocr_text, ocr_text, "paddleocr"
        except Exception as e:
            logger.error(f"  Paddle-OCR API failed: {e}. Falling back to local Tesseract OCR.")
            
    logger.info(f"  pdfplumber text too short ({len(bare)} chars) → local Tesseract OCR.")
    text_300 = extract_text_ocr(pdf_path, dpi=300)
    text_200 = extract_text_ocr(pdf_path, dpi=200)
    return text_300, text_200, "ocr"


def save_extracted_text(pdf_name: str, text: str, method: str) -> None:
    """Save raw extracted text to extracted_text/ for debugging."""
    out = EXTRACTED_TEXT_DIR / f"{Path(pdf_name).stem}__{method}.txt"
    out.write_text(text, encoding="utf-8")
    logger.debug(f"  Raw text saved → {out.name}")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 3: INVOICE SPLITTING (multi-invoice PDFs)
# ─────────────────────────────────────────────────────────────────────────────

def _find_inv_number(text: str) -> str:
    """Return first invoice number found in text, trying all patterns."""
    for pat in (_INV_NUM_PAT, _INV_NUM_PAT2):
        m = pat.search(text)
        if m:
            return m.group(1).strip()
    # Try bare invoice-number line pattern
    for line in text.splitlines():
        m = _INV_NUM_BARE.match(line)
        if m:
            return m.group(1).strip()
    return ""


def split_into_invoice_blocks_with_pages(text: str) -> list[list[tuple[int, str]]]:
    """
    Split text into list of blocks, where each block is a list of (page_index, page_text) tuples.
    page_index is 0-indexed.
    """
    pages = PAGE_SEP.split(text)
    pages = [p.strip() for p in pages]
    
    # Filter out empty pages but keep track of their original 0-indexed position
    indexed_pages = [(i, p) for i, p in enumerate(pages) if p.strip()]
    
    if not indexed_pages:
        return []
        
    blocks: list[list[tuple[int, str]]] = []
    current_block = [indexed_pages[0]]
    current_inv = _find_inv_number(indexed_pages[0][1])
    
    for idx, page in indexed_pages[1:]:
        starts_new = False
        this_inv = _find_inv_number(page)
        if this_inv and current_inv and this_inv != current_inv:
            starts_new = True
        else:
            lines = page.splitlines()
            header_idx = find_item_table_header(lines)
            if header_idx >= 0:
                row_start = re.compile(r"^\s*1\s*([|/\\-]\s*)?\s+\S")
                for k in range(header_idx + 1, len(lines)):
                    line = lines[k].strip()
                    if _FOOTER_PAT.match(line):
                        break
                    if row_start.match(lines[k]):
                        starts_new = True
                        break
                        
        if starts_new:
            blocks.append(current_block)
            current_block = [(idx, page)]
            current_inv = this_inv if this_inv else current_inv
        else:
            current_block.append((idx, page))
            if this_inv:
                current_inv = this_inv
                
    blocks.append(current_block)
    return blocks


def split_into_invoice_blocks(text: str) -> list[str]:
    """
    Split multi-page OCR/PDF text into separate invoice blocks.
    For backwards compatibility.
    """
    blocks_indexed = split_into_invoice_blocks_with_pages(text)
    return ["\n".join([p for _, p in block]) for block in blocks_indexed]


# ─────────────────────────────────────────────────────────────────────────────
# STEP 4: GENERIC HEADER EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def is_valid_po(val: str) -> bool:
    val_clean = val.strip()
    if len(val_clean) < 4:
        return False
    # If it has 3 or more lowercase letters in a row, it's likely a word
    if re.search(r"[a-z]{3,}", val_clean):
        return False
    # Exclude standard date patterns
    if re.search(r"\d{1,2}[\-/\.][A-Za-z0-9]+[\-/\.]\d{2,4}", val_clean):
        return False
    # Exclude common noise label words
    if val_clean.lower() in [
        "buyer", "consignee", "invoice", "dated", "date", "consigne", "page",
        "gstin", "state", "phone", "email", "total", "amount", "price", "quantity", "address", "name"
    ]:
        return False
    # Require at least one digit
    if not re.search(r"\d", val_clean):
        return False
    return True


def _extract_po_number(text: str) -> str:
    lines = text.splitlines()
    for i, line in enumerate(lines):
        m_lbl = _PO_LABELS.search(line)
        if m_lbl:
            # 1. Try same-line value first (only look after the matched label)
            after_label = line[m_lbl.end():].strip()
            for match in _PO_VALUE.finditer(after_label):
                val = match.group(1).strip()
                if is_valid_po(val):
                    return val
            # 2. Check lines after the label (up to 8 lines)
            for j in range(i + 1, min(i + 8, len(lines))):
                next_line = lines[j].strip()
                if next_line:
                    if any(x in next_line.lower() for x in ["gstin", "gst", "pan", "uin", "state", "address", "phone", "email"]):
                        continue
                    for match in _PO_VALUE.finditer(next_line):
                        val = match.group(1).strip()
                        if is_valid_po(val):
                            return val
            # 3. Check lines before the label (up to 5 lines)
            for j in range(max(0, i - 5), i):
                prev_line = lines[j].strip()
                if prev_line:
                    if any(x in prev_line.lower() for x in ["gstin", "gst", "pan", "uin", "state", "address", "phone", "email"]):
                        continue
                    for match in _PO_VALUE.finditer(prev_line):
                        val = match.group(1).strip()
                        if is_valid_po(val):
                            return val
    return ""


def _clean_invoice_number(raw: str) -> str:
    # Collapse internal spaces around slashes
    raw = re.sub(r"\s*/\s*", "/", raw)
    # Strip trailing noise words like Date, Dated, etc.
    raw = re.sub(
        r'\s+\b(Date|Dated|No\.?|Dt\.?|Time|Ref\.?|Number)\b\s*$',
        '', raw, flags=re.IGNORECASE
    ).strip()
    return raw


def _extract_invoice_number(text: str) -> str:
    lines = text.splitlines()
    # 1. Try primary labeled patterns
    for i, line in enumerate(lines[:50]):
        for pat in (_INV_NUM_PAT, _INV_NUM_PAT2):
            m = pat.search(line)
            if m:
                raw = _clean_invoice_number(m.group(1))
                # Filter out common labels to avoid incorrect extraction of headers/metadata
                if raw and raw.lower() not in ["date", "dated", "delivery", "note", "reference", "mode", "terms", "po", "buyer", "consignee", "total"]:
                    return raw
        # Also try matching label on line i, and value on line i+1
        m_lbl = re.search(r"\b(?:Invoice\s*(?:No\.?|Number)|Inv\.?\s*No\.?|Bill\s*No\.?|Tax\s*Invoice\s*No\.?|Invoice\s*No\s*:)", line, re.IGNORECASE)
        if m_lbl:
            after = line[m_lbl.end():].strip()
            if not after or after.lower() in ["dated", "date", "no", "no."]:
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    words = next_line.split()
                    if words:
                        cand = _clean_invoice_number(words[0])
                        if len(cand) >= 4 and any(c.isdigit() for c in cand):
                            return cand

    # 2. Try inline invoice patterns (uppercase only, like ATPL/26-27/00417)
    po_val = _extract_po_number(text)
    inline_pat = re.compile(r"\b([A-Z]{2,}[A-Z0-9/_-]*\d{2,})\b")
    
    ignore_lbls = [
        "gstin", "phone", "email", "buyer", "consignee", "address", "road", "street",
        "lane", "nagar", "cross", "marg", "bazar", "plaza", "complex", "building",
        "floor", "pincode", "zone", "state", "city", "district", "village", "taluk",
        "near", "opposite", "behind", "highway", "sector", "industrial", "park", "mall",
        "udyam", "pan", "bank", "a/c", "account", "ifsc", "rtgs", "neft", "tele", "fax",
        "website", "www.", "cin:"
    ]
    
    candidates = []
    for line in lines[:50]:
        if any(lbl in line.lower() for lbl in ignore_lbls):
            continue
        for match in inline_pat.finditer(line):
            val = _clean_invoice_number(match.group(1))
            if val and val != po_val and len(val) >= 4:
                candidates.append(val)
                
    if candidates:
        return candidates[0]

    # 3. Bare pattern on own line
    for line in lines[:40]:
        m = _INV_NUM_BARE.match(line)
        if m:
            return _clean_invoice_number(m.group(1))

    return ""


def normalize_date(raw_date: str) -> str:
    d = raw_date.strip()
    if not d:
        return ""
    # Replace separators with hyphens or slashes for easy parsing
    d = re.sub(r"[\.\s]+", "-", d)
    d = re.sub(r"/", "-", d)
    
    parts = d.split("-")
    if len(parts) != 3:
        return raw_date
        
    day_str, month_str, year_str = parts[0], parts[1], parts[2]
    
    day = re.sub(r"\D", "", day_str)
    if not day:
        return raw_date
    day_int = int(day)
    
    month_str_clean = month_str.strip().lower()
    months_map = {
        "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
        "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
        "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
        "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12
    }
    if month_str_clean in months_map:
        month_int = months_map[month_str_clean]
    else:
        month_digits = re.sub(r"\D", "", month_str)
        if not month_digits:
            return raw_date
        month_int = int(month_digits)
        
    year = re.sub(r"\D", "", year_str)
    if not year:
        return raw_date
    if len(year) == 2:
        year_int = 2000 + int(year)
    else:
        year_int = int(year)
        
    if 1 <= day_int <= 31 and 1 <= month_int <= 12 and 1900 <= year_int <= 2100:
        return f"{day_int:02d}/{month_int:02d}/{year_int:4d}"
        
    return raw_date


def _extract_invoice_date(text: str) -> str:
    """Extract invoice date, also checking line after label."""
    # 1. Try to find date on the same line as the invoice number
    inv_num = _extract_invoice_number(text)
    if inv_num:
        for line in text.splitlines():
            if inv_num in line:
                # Search for a date pattern on this line
                date_val = re.compile(
                    r"(\d{1,2}[\-/\.][A-Za-z0-9]+[\-/\.]\d{2,4}"
                    r"|\d{1,2}[\-/\.][A-Za-z0-9]+[\-/\.]\d{2}"
                    r"|\d{1,2}/\d{1,2}/\d{2,4}"
                    r"|\d{1,2}-\d{1,2}-\d{2,4})"
                )
                for m in date_val.finditer(line):
                    date_str = m.group(1).strip()
                    if date_str not in inv_num:
                        return normalize_date(date_str)

    # 2. Try primary labeled patterns
    m = _INV_DATE_PAT.search(text)
    if m:
        return normalize_date(m.group(1).strip())

    # 3. Try: label on one line, date on next
    lines = text.splitlines()
    date_label = re.compile(
        r"Invoice\s*Date|Inv\.?\s*Date|Bill\s*Date|Document\s*Date|Dated",
        re.IGNORECASE,
    )
    date_val = re.compile(
        r"(\d{1,2}[\-/\.][A-Za-z0-9]+[\-/\.]\d{2,4}|\d{1,2}\s+[A-Za-z]+\s+\d{4})"
    )
    for i, line in enumerate(lines):
        if date_label.search(line):
            # Try next few lines
            for j in range(i, min(i + 4, len(lines))):
                dm = date_val.search(lines[j])
                if dm:
                    return normalize_date(dm.group(1).strip())
    return ""


def _extract_vendor_name(text: str) -> str:
    """
    Extract vendor/company name from the invoice text.
    Uses multiple fallback strategies:
      1. GSTIN/PAN matching
      2. Signature block detection ("For <Vendor>")
      3. Header pattern matching (first 40 lines)
    """
    # 1. Signature block matching ("For <Vendor>")
    for line in text.splitlines():
        m_for = re.search(r"\bFor\s+([A-Za-z0-9&.'\- ]{3,60})", line)
        if m_for:
            candidate = m_for.group(1).strip()
            # Exclude buyer and noise labels
            exclude_words = [
                "skaps", "customer", "buyer", "consignee", "signature", "us", "you",
                "recipient", "purchaser", "shipper", "receiver", "client", "vendor",
                "seller", "supplier", "party", "above", "same", "behalf", "directors",
                "office", "bank", "attention", "attn", "detail", "details"
            ]
            if not any(x in candidate.lower() for x in exclude_words):
                candidate = re.sub(r"\b(?:Authorized|Authorised|Signatory|Signature|Director|Partner|Proprietor)\b", "", candidate, flags=re.IGNORECASE).strip()
                candidate = re.sub(r"\s*[|/\\\-]$", "", candidate).strip()
                if len(candidate) > 3:
                    return candidate

    # 2. Header-based company names (first 40 lines)
    header_lines = text.splitlines()[:40]
    header_text = "\n".join(header_lines)
    matches = _VENDOR_KEYWORDS.findall(header_text)
    if matches:
        for match in matches:
            name = match.strip()
            name = re.sub(r"\s{2,}", " ", name)
            words = name.split()
            cleaned_words = []
            started = False
            for word in words:
                if not started:
                    if re.search(r"\d", word) or word.lower() in ["dated", "date", "no", "no.", "invoice"]:
                        continue
                    started = True
                cleaned_words.append(word)
            if cleaned_words:
                name = " ".join(cleaned_words)
            if name and "skaps" not in name.lower():
                return name
                
    return ""


def extract_header_generic(text: str) -> dict:
    """
    Extract Invoice Date, Invoice Number, PO Number, and Vendor Name
    using broad patterns covering many Indian vendor invoice layouts.
    """
    return {
        "Invoice Date":    _extract_invoice_date(text),
        "Invoice Number":  _extract_invoice_number(text),
        "PO Number":       _extract_po_number(text),
        "Vendor Name":     _extract_vendor_name(text),
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 5: HSN EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def extract_hsn_from_block(block_text: str) -> str:
    """
    Extract HSN / SAC code from an invoice block.
    Tries explicit label first ('HSN Code : 997331'), then inline 4-8 digit numbers.
    Returns alphanumeric HSN code or empty string.
    """
    pat = re.compile(
        r"(?:HSN\s*/\s*SAC|HSN\s*(?:Code)?|SAC\s*(?:Code)?)"
        r"[\s\n\-\|:\/]*"
        r"\b(\d{4,10})\b",
        re.IGNORECASE
    )
    m = pat.search(block_text)
    if m:
        return m.group(1).strip()
    return ""


def extract_all_hsns(block_text: str) -> list[str]:
    """Extract all HSN codes from a block text."""
    pat = re.compile(
        r"(?:HSN\s*/\s*SAC|HSN\s*(?:Code)?|SAC\s*(?:Code)?)"
        r"[\s\n\-\|:\/]*"
        r"\b(\d{4,10})\b",
        re.IGNORECASE
    )
    hsns = []
    for m in pat.finditer(block_text):
        h = m.group(1).strip()
        if h not in hsns:
            hsns.append(h)
    
    if hsns:
        return hsns
        
    for m in _HSN_PAT.finditer(block_text):
        val = m.group(1)
        if val in ["2024", "2025", "2026", "2027", "382330", "370421"]:
            continue
        if val not in hsns:
            hsns.append(val)
    return hsns


def extract_hsn_from_line(raw: str) -> str:
    """
    Extract HSN from an item table row (inline numeric code).
    Returns a 4-8 digit code or ''.
    """
    m = _HSN_ALPHA_PAT.search(raw)
    if m:
        return m.group(1).strip()
    # Fallback: first 4-8 digit standalone number
    m2 = _HSN_PAT.search(raw)
    if m2:
        return m2.group(1)
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# STEP 6: ITEM TABLE DETECTION & PARSING
# ─────────────────────────────────────────────────────────────────────────────

def find_item_table_header(lines: list[str]) -> int:
    """
    Return the index of the table header row (containing item + qty/amount
    column headings), or -1 if not found.
    """
    for i, line in enumerate(lines):
        if _TH_ITEM.search(line) and (_TH_QTY.search(line) or _TH_AMOUNT.search(line)):
            return i
    return -1


def detect_pipe_columns(header_line: str) -> dict:
    """
    If the header line is pipe-separated, detect which column index corresponds
    to Description, HSN, Quantity, Rate, and Amount.
    Returns a dict with keys: desc, hsn, qty, rate, amount (values are col indices or -1).
    """
    if '|' not in header_line:
        return {}
    cols = [c.strip().lower() for c in header_line.split('|')]
    result = {'desc': -1, 'hsn': -1, 'qty': -1, 'rate': -1, 'amount': -1}
    for i, col in enumerate(cols):
        if re.search(r'description|particulars|item|name\s*of', col):
            result['desc'] = i
        elif re.search(r'hsn|sac|hsncode', col):
            result['hsn'] = i
        elif re.search(r'\bqty\b|\bquantity\b|\bqnty\b|\bnos\b|\bunits?\b|\bpcs\b', col):
            result['qty'] = i
        elif re.search(r'\brate\b|\bprice\b|\bunit\s*rate\b', col):
            result['rate'] = i
        elif re.search(r'\bamount\b|\btaxable\b|\btotal\b', col):
            result['amount'] = i
    # Return only if we found at least description + one numeric column
    if result['desc'] >= 0 and (result['qty'] >= 0 or result['amount'] >= 0):
        return result
    return {}


def extract_from_pipe_row(line: str, col_map: dict) -> tuple[str, str, str, str]:
    """
    Given a pipe-separated data row and a column map from detect_pipe_columns(),
    extract (description, hsn, qty, price) directly by column index.
    Returns ('', '', '', '') if the row doesn't have enough columns.
    """
    parts = [p.strip() for p in line.split('|')]
    max_needed = max(v for v in col_map.values() if v >= 0)
    if len(parts) <= max_needed:
        return '', '', '', ''

    desc = parts[col_map['desc']].strip() if col_map.get('desc', -1) >= 0 else ''
    hsn  = parts[col_map['hsn']].strip()  if col_map.get('hsn', -1)  >= 0 else ''
    qty_raw    = parts[col_map['qty']].strip()    if col_map.get('qty', -1)    >= 0 else ''
    amount_raw = parts[col_map['amount']].strip() if col_map.get('amount', -1) >= 0 else ''
    rate_raw   = parts[col_map['rate']].strip()   if col_map.get('rate', -1)   >= 0 else ''

    # Clean qty: strip reading numbers, dates, percentages — keep first plain number
    qty = ''
    qty_f = 0.0
    if qty_raw:
        tmp = _strip_dates(qty_raw)
        tmp = re.sub(r'\b\d+(?:\.\d+)?\s*%', ' ', tmp)
        tmp = re.sub(r'\b(?:Last|Current|Opening|Closing|Meter)?\s*Reading\s*\d+\b', ' ', tmp, flags=re.IGNORECASE)
        for m in re.finditer(r'\b(\d{1,6}(?:\.\d{1,3})?)\b', tmp):
            v = m.group(1)
            if not _is_year(v):
                qty = v
                try:
                    qty_f = float(v)
                except ValueError:
                    pass
                break

    # Clean rate: extract first plain number from Rate column
    rate_f = 0.0
    if rate_raw:
        tmp_r = re.sub(r'\b\d+(?:\.\d+)?\s*%', ' ', rate_raw)
        for m in re.finditer(r'\b(\d{1,6}(?:\.\d{1,3})?)\b', tmp_r):
            v = m.group(1)
            if not _is_year(v):
                try:
                    rate_f = float(v)
                except ValueError:
                    pass
                break

    # Clean amount: take first monetary-looking value from Amount column
    price = ''
    if amount_raw:
        tmp = re.sub(r'\\n.*', '', amount_raw)  # strip escaped newlines
        money_m = _MONEY_PAT.search(tmp)
        if money_m:
            price = money_m.group(0)
        else:
            num_m = re.search(r'\b(\d+(?:\.\d+)?)\b', tmp)
            if num_m and not _is_year(num_m.group(1)):
                price = num_m.group(1)

    # When no Amount column exists, compute price = qty x rate
    if not price and qty_f > 0 and rate_f > 0:
        computed = qty_f * rate_f
        price = f"{computed:.2f}"

    # Clean description: strip leading row number, truncate at reading/model noise
    desc = re.sub(r'^\s*\d{1,3}\s*', '', desc).strip()
    m_trunc = _DESC_TRUNCATE_AT.search(desc)
    if m_trunc:
        desc = desc[:m_trunc.start()].strip()
    # Also strip inline reading patterns that weren't caught by _DESC_TRUNCATE_AT
    desc = re.sub(
        r'\s*(?:Last|Current|Opening|Closing)\s+Reading\s+\d+.*$', '',
        desc, flags=re.IGNORECASE
    ).strip()
    desc = re.sub(r'\s{2,}', ' ', desc).strip()

    # Clean HSN: take first 4-8 digit number
    if hsn:
        hsn_m = re.search(r'\b(\d{4,8})\b', hsn)
        hsn = hsn_m.group(1) if hsn_m else ''

    return desc, hsn, qty, price



def parse_numbers_from_line(line: str) -> list[str]:
    """Return all monetary values found in a line (e.g. '13,000.00')."""
    return _MONEY_PAT.findall(line)


def _clean_description(raw: str) -> str:
    """Remove HSN codes, monetary values, units, row numbers, and pipe chars from a description."""
    d = raw
    d = _MONEY_PAT.sub(" ", d)        # remove prices
    # Only remove standalone HSN-like digit blocks (not part of product codes like "842315 TONER")
    d = re.sub(r'(?<!\w)\d{4,8}(?!\s*[A-Za-z])', ' ', d)
    d = _UNIT_PAT.sub(" ", d)         # remove unit tokens (NOS, PCS, etc.)
    d = re.sub(r"\|", " ", d)         # remove pipe separators
    d = re.sub(r"^\s*\d{1,3}\s*", "", d)  # remove leading row number
    d = re.sub(r"\s{2,}", " ", d)
    return d.strip()


def extract_description_and_hsn(raw: str) -> tuple[str, str]:
    """
    From a raw item line (after the row-number prefix is removed),
    split out the text description and numeric HSN code.
    Returns (description, hsn).
    """
    # Strip dates and reading numbers first to avoid matching them as HSN/description
    raw_clean = _strip_dates(raw)
    raw_clean = re.sub(r"\b(?:Last|Current|Opening|Closing|Meter)?\s*Reading\s*\d+\b", " ", raw_clean, flags=re.IGNORECASE)
    raw_clean = re.sub(r"\bReading\b.*?\b\d+\b", " ", raw_clean, flags=re.IGNORECASE)
    
    # Try to find the standalone HSN (4-8 consecutive digits, not followed by word characters) position
    hsn_m = re.search(r"\b(\d{4,8})\b(?!\s*[A-Za-z])", raw_clean)
    if hsn_m:
        hsn_val = hsn_m.group(1)
        # Find where hsn_val is in the original raw string
        idx = raw.find(hsn_val)
        if idx != -1:
            # If HSN is at the beginning of raw (e.g. index < 15), description is after it
            if idx < 15 and len(raw[idx + len(hsn_val):].strip()) > 3:
                description = raw[idx + len(hsn_val):].strip()
            else:
                description = raw[:idx].strip()
            description = re.sub(r"[|]", " ", description)
            description = re.sub(r"\s{2,}", " ", description)
            return description.strip(), hsn_val

    # No HSN – description is everything up to first standalone integer/decimal column
    num_m = _STANDALONE_NUM_PAT.search(raw)
    if num_m:
        matches = list(_STANDALONE_NUM_PAT.finditer(raw))
        if matches:
            money_matches = list(_MONEY_PAT.finditer(raw))
            if money_matches:
                cutoff = money_matches[0].start()
                for m in reversed(matches):
                    if m.start() < cutoff:
                        if cutoff - m.end() < 5:
                            cutoff = m.start()
                description = raw[:cutoff].strip()
                description = re.sub(r"[|]", " ", description)
                description = re.sub(r"\s{2,}", " ", description)
                return description.strip(), ""
        
        description = raw[:num_m.start()].strip()
        description = re.sub(r"[|]", " ", description)
        description = re.sub(r"\s{2,}", " ", description)
        return description.strip(), ""

    return raw.strip(), ""


def _strip_dates(text: str) -> str:
    """Remove date-like components from a string to prevent them from matching as quantities/prices/HSNs."""
    # Remove standard DD-MM-YYYY or YYYY-MM-DD dates (with dashes, slashes, or dots)
    text = re.sub(r"\b\d{1,4}[-/\.]\d{1,2}[-/\.]\d{2,4}\b", " ", text)
    # Remove dates like "09 Apr 2025" or "09-Apr-2025"
    text = re.sub(r"\b\d{1,2}[-\s]+[A-Za-z]{3,}[-\s]+\d{2,4}\b", " ", text, flags=re.IGNORECASE)
    return text


def extract_qty(raw: str) -> str:
    """
    Extract quantity from the numeric part of an item line.
    Strategy:
      1. Try matching a number followed by a unit of measure (most reliable).
      2. Fallback: remove only large monetary values (with commas) and HSN/units to avoid removing small quantities.
    """
    # Strip leading row number prefix
    raw_clean = re.sub(r"^\s*\d{1,3}\s*(?:[|/\\-]\s*|[IiLl]\s+)?", "", raw)
    
    qty_unit_pat = re.compile(
        r"\b(\d+(?:\.\d+)?)\s*(?:[|]?\s*)*(?:NOS|PCS|UNIT|UNITS|EA|SET|KG|MTR|LTR|BOX|PC|ROLL|EACH|NO\.)\b",
        re.IGNORECASE
    )
    m = qty_unit_pat.search(raw_clean)
    if m:
        return m.group(1)

    temp = _strip_dates(raw_clean)
    # Strip percentage values (like 18.00% or 18 %)
    temp = re.sub(r"\b\d+(?:\.\d+)?\s*%", " ", temp)
    temp = re.sub(r"\b(?:Last|Current|Opening|Closing|Meter)?\s*Reading\s*\d+\b", " ", temp, flags=re.IGNORECASE)
    temp = re.sub(r"\bReading\b.*?\b\d+\b", " ", temp, flags=re.IGNORECASE)
    temp = re.sub(r"\b\d{1,3}(?:,\d{3})+\.\d{1,2}\b", " ", temp) # remove comma prices
    temp = _HSN_PAT.sub(" ", temp)        # remove HSN
    temp = _UNIT_PAT.sub(" ", temp)       # remove unit tokens
    # First remaining standalone integer or simple decimal that is not a year
    for match in re.finditer(r"\b(\d{1,5}(?:\.\d{1,3})?)\b", temp):
        val = match.group(1)
        if not _is_year(val):
            return val
    return ""


def _is_year(v_str: str) -> bool:
    """Return True if the value string looks like a calendar year (1900-2100)."""
    try:
        val_f = float(v_str.replace(",", ""))
        return 1900 <= val_f <= 2100 and val_f == int(val_f)
    except ValueError:
        return False


def _is_pipe_heavy(line: str) -> bool:
    """Return True if the line contains more than 3 vertical pipes."""
    return line.count("|") > 3


def _select_price(money_vals: list[str], qty_str: str) -> str:
    """
    From a list of monetary values on an item line, select the most likely
    'Basic Price excluding GST' (taxable / unit value).

    Rules:
    - If qty is known, prefer the value that equals rate × qty or is the
      largest pre-tax value (ignoring tax % values like 9.00, 18.00).
    - Otherwise return the largest monetary value that is NOT a tiny tax rate.
    """
    # Filter out obvious tax-rate-like values (single/double digit decimals like 9.00, 18.00)
    # and also exclude year-like values (like 2025, 2026)
    filtered = [v for v in money_vals if float(v.replace(",", "")) > 99 and not _is_year(v)]
    if not filtered:
        filtered = [v for v in money_vals if not _is_year(v)]
    if not filtered:
        filtered = money_vals
    if not filtered:
        return ""

    # Check if one value is the after-tax version of another (e.g. at 18%, 12%, 5%, 28%)
    # If so, the smaller one is the taxable basic price!
    tax_multipliers = [1.18, 1.12, 1.05, 1.28, 1.09, 1.15, 1.025]
    for v1 in filtered:
        try:
            v1f = float(v1.replace(",", ""))
            for v2 in filtered:
                v2f = float(v2.replace(",", ""))
                if v2f > v1f:
                    for mult in tax_multipliers:
                        if abs(v1f * mult - v2f) < 2.0:
                            return v1
        except ValueError:
            pass

    # Try to find taxable value = qty * rate
    if qty_str:
        try:
            qty_f = float(qty_str)
            for v in filtered:
                vf = float(v.replace(",", ""))
                # Check if any other value = this value * qty (implies this is the unit rate)
                # or if this value / qty gives a clean rate
                if qty_f > 0:
                    unit_rate = vf / qty_f
                    # If a value in the list matches vf (taxable = qty * rate), prefer it
                    for v2 in filtered:
                        v2f = float(v2.replace(",", ""))
                        if abs(v2f - unit_rate * qty_f) < 1:
                            return v2
        except (ValueError, ZeroDivisionError):
            pass

    # Fallback: return the largest value (total taxable amount)
    try:
        return max(filtered, key=lambda x: float(x.replace(",", "")))
    except ValueError:
        return filtered[0]


def _collect_continuation_lines(lines: list[str], start_idx: int,
                                 max_lines: int = 5) -> str:
    """
    After a primary item row, collect continuation description lines.
    Stops at: footer pattern, noise, a new numbered row, empty line, or pipe-heavy line.
    """
    extra_parts = []
    row_num_re = re.compile(r"^\s*\d{1,3}\s*[|IiLl]?\s+\S")
    for j in range(start_idx, min(start_idx + max_lines, len(lines))):
        line = lines[j].strip()
        if not line:
            break
        if _FOOTER_PAT.match(line):
            break
        if row_num_re.match(lines[j]):
            break          # new item row starts
        if looks_like_noise(line) or _is_pipe_heavy(line):
            break
        # Exclude HSN codes from the continuation text to keep the Item name clean
        if "hsn" in line.lower():
            continue
        # Only add if it looks like a meaningful text continuation
        # (not all numbers / codes)
        if re.search(r"[A-Za-z]{3,}", line):
            extra_parts.append(line)
        else:
            break
    return " ".join(extra_parts)


def parse_qty_and_price(line: str, hsn_code: str, row_num: str, money_vals: list[str] = None) -> tuple[str, str, str]:
    num_pat = re.compile(r"\b\d+(?:,\d{3})*(?:\.\d+)?\b")
    
    matches = list(num_pat.finditer(line))
    numbers = []
    for idx, match in enumerate(matches):
        m_str = match.group(0)
        val_str = m_str.replace(",", "")
        
        if hsn_code and m_str == hsn_code:
            continue
            
        # Check if match is part of a date or reading
        is_ignored = False
        for date_match in re.finditer(r"\b\d{1,4}[-/\.]\d{1,2}[-/\.]\d{2,4}\b", line):
            if date_match.start() <= match.start() < date_match.end():
                is_ignored = True
                break
        if not is_ignored:
            for date_match in re.finditer(r"\b\d{1,2}[-\s]+[A-Za-z]{3,}[-\s]+\d{2,4}\b", line, flags=re.IGNORECASE):
                if date_match.start() <= match.start() < date_match.end():
                    is_ignored = True
                    break
        if not is_ignored:
            for reading_match in re.finditer(r"\b(?:Last|Current|Opening|Closing|Meter)?\s*Reading\s*\d+\b", line, flags=re.IGNORECASE):
                if reading_match.start() <= match.start() < reading_match.end():
                    is_ignored = True
                    break
        if not is_ignored:
            for reading_match in re.finditer(r"\bReading\b.*?\b\d+\b", line, flags=re.IGNORECASE):
                m_num = re.search(r"\b\d+\b", reading_match.group(0))
                if m_num:
                    num_start = reading_match.start() + m_num.start()
                    num_end = reading_match.start() + m_num.end()
                    if num_start <= match.start() < num_end:
                        is_ignored = True
                        break
                        
        if is_ignored:
            continue
            
        try:
            val_f = float(val_str)
            if not _is_year(m_str):
                numbers.append((idx, m_str, val_f, match.start()))
        except ValueError:
            pass
            
    hsn_idx = len(line)
    if hsn_code:
        hsn_m = re.search(re.escape(hsn_code), line)
        if hsn_m:
            hsn_idx = hsn_m.start()
            
    triplets = []
    for i1, str1, f1, start1 in numbers:
        if f1 < 0.5:
            continue
        for i2, str2, f2, start2 in numbers:
            if i2 == i1 or f2 < 0.5:
                continue
            for i3, str3, f3, start3 in numbers:
                if i3 == i1 or i3 == i2 or f3 < 0.5:
                    continue
                if abs(f1 * f2 - f3) < 1.5:
                    # Domain rule: if one factor is a fraction (< 1) and the other is
                    # a whole number, the whole number is the quantity (e.g. 7288 copies
                    # at ₹0.38 each). Using min/max would incorrectly assign qty=0.38.
                    domain_rule_applied = False
                    if f1 < 1.0 and f2 >= 1.0 and f2 == int(f2):
                        qty, rate = f2, f1
                        domain_rule_applied = True
                    elif f2 < 1.0 and f1 >= 1.0 and f1 == int(f1):
                        qty, rate = f1, f2
                        domain_rule_applied = True
                    else:
                        qty = min(f1, f2)
                        rate = max(f1, f2)
                    # Refine using extract_qty hint — but only if it gives a sensible qty
                    # (>= 1) or if no domain rule was applied. Never let a fractional hint
                    # override a domain-rule whole-number quantity.
                    qty_est = extract_qty(line)
                    if qty_est:
                        try:
                            qty_est_f = float(qty_est)
                            hint_sensible = qty_est_f >= 1.0
                            if not domain_rule_applied or hint_sensible:
                                if abs(f1 - qty_est_f) < 0.01:
                                    qty = f1
                                    rate = f2
                                elif abs(f2 - qty_est_f) < 0.01:
                                    qty = f2
                                    rate = f1
                        except ValueError:
                            pass
                    qty_idx = start1 if f1 == qty else start2
                    rate_idx = start2 if f1 == qty else start1
                    triplets.append((qty, rate, f3, str3, qty_idx, rate_idx, start3))
                    
    m_prefix = re.match(r"^\s*\d{1,3}\s*(?:[|/\\-]\s*|[IiLl]\s+)?", line)
    prefix_len = m_prefix.end() if m_prefix else 0
    
    if triplets:
        triplets.sort(key=lambda x: (-x[2], -x[0], -x[4]))
        best_qty, best_rate, best_price, price_str, qty_idx, rate_idx, amount_idx = triplets[0]
        
        qty_str = f"{best_qty:.2f}" if best_qty % 1 != 0 else f"{int(best_qty)}"
        # Price is now the total amount before tax (best_price)
        price_str = f"{best_price:,.2f}" if best_price % 1 != 0 else f"{int(best_price):,}"
        
        for idx, m_str, val_f, start in numbers:
            if abs(val_f - best_price) < 0.01 and start != 0:
                price_str = m_str
                break
                
        start_desc = prefix_len
        if hsn_code and hsn_idx < prefix_len + 15:
            start_desc = hsn_idx + len(hsn_code)
            
        cutoff = min(rate_idx, amount_idx)
        if qty_idx > start_desc:
            cutoff = min(cutoff, qty_idx)
        if hsn_code and hsn_idx > start_desc:
            cutoff = min(cutoff, hsn_idx)
            
        desc = line[start_desc:cutoff].strip()
        desc = re.sub(r"\s*[|IiLl/\\-]$", "", desc).strip()
        return qty_str, price_str, desc
        
    # Fallback
    fallback_money = money_vals if money_vals is not None else parse_numbers_from_line(line)
    qty_str = extract_qty(line)
    price_str = _select_price(fallback_money, qty_str)
    desc, hsn = extract_description_and_hsn(line)
    desc = re.sub(r"\s*[|IiLl/\\-]$", "", desc).strip()
    return qty_str, price_str, desc


def extract_items_from_table(lines: list[str], header_idx: int) -> list[dict]:
    """
    Extract item rows from lines below the table header.
    Each row must:
      - Not be noise.
      - Start with a row number (1-3 digits).
      - Contain at least one monetary value.
    Continuation description lines (word-only, no new row number) are appended
    to the preceding item's description.
    Returns list of dicts: {item, hsn, qty, price, line_idx}.
    """
    items: list[dict] = []
    # Matches: optional spaces, 1-3 digit row number, optional | / I separator, content
    row_start = re.compile(r"^\s*(\d{1,3})\s*[|IiLl]?\s+(.+)", re.DOTALL)

    # Detect pipe-table column layout from header (for Paddle-OCR HTML-converted tables)
    col_map = detect_pipe_columns(lines[header_idx]) if header_idx >= 0 else {}

    i = header_idx + 1
    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        if _FOOTER_PAT.match(stripped):
            break
        if looks_like_noise(stripped):
            i += 1
            continue

        m = row_start.match(line)
        if not m:
            i += 1
            continue

        raw_content = m.group(2).strip()
        money_vals = parse_numbers_from_line(raw_content)

        # Check next 2 lines for monetary values if none found or if they are all small
        has_large_money = any(float(v.replace(",", "")) > 99 and not _is_year(v) for v in money_vals)
        if not has_large_money:
            for offset in range(1, 3):
                if i + offset < len(lines):
                    next_line = lines[i + offset]
                    if row_start.match(next_line) or _FOOTER_PAT.match(next_line.strip()):
                        break
                    next_money = parse_numbers_from_line(next_line)
                    if next_money:
                        money_vals.extend(next_money)

        if not money_vals:
            i += 1
            continue                     # no monetary value → skip

        # ── Pipe-table fast path ────────────────────────────────────────────
        # When the header has labeled pipe columns, read qty/price directly
        # from the correct column — bypasses regex confusion with reading nums.
        description, hsn, qty, price = '', '', '', ''
        used_pipe_path = False
        if col_map and '|' in line:
            description, hsn, qty, price = extract_from_pipe_row(line, col_map)
            if description or qty or price:
                used_pipe_path = True

        if not used_pipe_path:
            description, hsn = extract_description_and_hsn(raw_content)
            qty, price, parsed_desc = parse_qty_and_price(line, hsn, m.group(1).strip(), money_vals)
            if parsed_desc:
                description = parsed_desc

        # Collect continuation lines for multi-line descriptions
        cont = _collect_continuation_lines(lines, i + 1, max_lines=4)
        if cont:
            description = (description + " " + cont).strip()
            description = re.sub(r"\s{2,}", " ", description)

        # Truncate at noise headers (e.g. Contract Period, Model No, etc.)
        m_trunc = _DESC_TRUNCATE_AT.search(description)
        if m_trunc:
            description = description[:m_trunc.start()].strip()

        description = re.sub(r"\s*[|IiLl/\\-]$", "", description).strip()

        # Skip very short or obviously non-item descriptions
        words = [w for w in description.split() if len(w) > 1 and w.isascii()]
        if len(words) < 2:
            logger.debug(f"  Skipped (too few words): '{description[:60]}'")
            i += 1
            continue
        if looks_like_noise(description):
            logger.debug(f"  Skipped (noise desc): '{description[:60]}'")
            i += 1
            continue

        # If HSN not found inline, try next few lines (some vendors put it below)
        if not hsn:
            for k in range(i + 1, min(i + 4, len(lines))):
                hsn_m = _HSN_ALPHA_PAT.search(lines[k])
                if hsn_m:
                    hsn = hsn_m.group(1).strip()
                    break
                # Also try plain numeric HSN on dedicated "HSN Code : XXXX" line
                plain_m = re.search(r"HSN\s*(?:Code)?\s*[:\-]?\s*(\d{4,8})", lines[k], re.IGNORECASE)
                if plain_m:
                    hsn = plain_m.group(1)
                    break

        items.append({
            "item":     description,
            "hsn":      hsn,
            "qty":      qty,
            "price":    price,
            "line_idx": i,
        })
        i += 1

    return items


# ─────────────────────────────────────────────────────────────────────────────
# STEP 7: MAKE A RECORD
# ─────────────────────────────────────────────────────────────────────────────

def make_record(
    source_file: str,
    header: dict,
    item: str = "",
    hsn: str = "",
    qty: str = "",
    price: str = "",
    status: str = "Success",
    remarks: str = "",
) -> dict:
    """Assemble one flat row for the Excel output."""
    missing = []
    if not header.get("Invoice Date"):
        missing.append("Invoice Date")
    if not header.get("Invoice Number"):
        missing.append("Invoice Number")
    if not item:
        missing.append("Item")
    if not qty:
        missing.append("Quantity")
    if not price:
        missing.append("Basic Price")

    if missing:
        status = "Partial" if status == "Success" else status
        note   = f"Missing: {', '.join(missing)}"
        remarks = f"{remarks}; {note}".strip("; ") if remarks else note

    return {
        "Source File":               source_file,
        "Vendor Name":               header.get("Vendor Name", ""),
        "Invoice Date":              header.get("Invoice Date", ""),
        "Invoice Number":            header.get("Invoice Number", ""),
        "PO Number":                 header.get("PO Number", ""),
        "Item":                      item,
        "HSN Number":                hsn,
        "Quantity":                  qty,
        "Basic Price excluding GST": price,
        "Extraction Status":         status,
        "Remarks":                   remarks,
    }


# ─────────────────────────────────────────────────────────────────────────────
# STEP 8: GENERIC ITEM EXTRACTOR
# ─────────────────────────────────────────────────────────────────────────────

def infer_qty_and_rate(money_vals: list[str]) -> tuple[str, str]:
    """
    If we have two money values and one is a multiple of the other (with a small integer multiplier, e.g. between 1 and 1000):
    We can infer that the smaller one is the Unit Rate, the larger one is the Total Amount, and the multiplier is the Quantity!
    """
    if len(money_vals) >= 2:
        floats = []
        for mv in money_vals:
            try:
                floats.append((mv, float(mv.replace(",", ""))))
            except ValueError:
                pass
        floats.sort(key=lambda x: x[1], reverse=True)
        for i in range(len(floats)):
            total_str, total_val = floats[i]
            for j in range(i + 1, len(floats)):
                rate_str, rate_val = floats[j]
                if rate_val > 0:
                    ratio = total_val / rate_val
                    if abs(ratio - round(ratio)) < 0.05 and 1 <= round(ratio) <= 1000:
                        qty = round(ratio)
                        return str(qty), total_str
    return "", ""


def extract_items_fallback(lines: list[str]) -> list[dict]:
    """
    Fallback line-by-line item extraction when no table is detected.
    Looks for lines containing alphabetical descriptions and numeric quantities/prices.
    """
    items = []
    
    for i, line in enumerate(lines):
        stripped = line.strip()
        if not stripped:
            continue
        if looks_like_noise(stripped) or _is_pipe_heavy(stripped):
            continue
        if _FOOTER_PAT.match(stripped):
            continue
            
        money_vals = parse_numbers_from_line(line)
        if not money_vals:
            continue
            
        description, hsn = extract_description_and_hsn(line)
        
        # Truncate description at noise keywords
        m_trunc = _DESC_TRUNCATE_AT.search(description)
        if m_trunc:
            description = description[:m_trunc.start()].strip()
        description = re.sub(r"\s*[|IiLl/\\-]$", "", description).strip()
        
        desc_clean = _clean_description(description)
        words = [w for w in desc_clean.split() if len(w) > 1 and w.isascii()]
        if len(words) < 2:
            continue
            
        qty, price = infer_qty_and_rate(money_vals)
        if not qty:
            qty, price, parsed_desc = parse_qty_and_price(line, hsn, "")
            if parsed_desc:
                description = parsed_desc
            if not qty:
                qty = "1"
            if not price:
                price = _select_price(money_vals, qty)
                
        # Re-apply truncation if description changed during parsing
        m_trunc = _DESC_TRUNCATE_AT.search(description)
        if m_trunc:
            description = description[:m_trunc.start()].strip()
        description = re.sub(r"\s*[|IiLl/\\-]$", "", description).strip()
        
        if not hsn:
            for k in range(i + 1, min(i + 4, len(lines))):
                hsn_m = _HSN_ALPHA_PAT.search(lines[k])
                if hsn_m:
                    hsn = hsn_m.group(1).strip()
                    break
                plain_m = re.search(r"HSN\s*(?:Code)?\s*[:\-]?\s*(\d{4,8})", lines[k], re.IGNORECASE)
                if plain_m:
                    hsn = plain_m.group(1)
                    break
                    
        items.append({
            "item":     description,
            "hsn":      hsn,
            "qty":      qty,
            "price":    price,
            "line_idx": i,
        })
        
    return items


def extract_generic_items(
    block_text: str,
    block_fallback: str,
    source_file: str,
    method: str,
    valid_vendor: str = "",
) -> list[dict]:
    """
    Generic item extractor for one invoice block.

    Workflow:
    1. Extract header fields (Invoice Date, Number, PO, Vendor Name).
    2. Find the item table header row.
    3. Extract rows below it using strict noise filtering.
    4. Collect HSN codes per item (inline or nearby lines).
    5. If no table found → create ONE Partial row.
    """
    header = extract_header_generic(block_text)

    # Merge fallback header fields from 200 DPI OCR text
    if block_fallback.strip():
        header_fallback = extract_header_generic(block_fallback)
        for key in ["Invoice Number", "Invoice Date", "PO Number", "Vendor Name"]:
            val_orig = header.get(key, "").strip()
            val_fallback = header_fallback.get(key, "").strip()
            if not val_orig and val_fallback:
                logger.debug(f"  [Fallback] Filled empty {key} with '{val_fallback}'")
                header[key] = val_fallback
            elif val_orig and val_fallback and key == "Invoice Number" and len(val_orig) < len(val_fallback) and val_fallback.startswith(val_orig):
                logger.debug(f"  [Fallback] Replaced short {key} '{val_orig}' with '{val_fallback}'")
                header[key] = val_fallback

    # Apply valid vendor inheritance
    v_orig = header.get("Vendor Name", "").strip()
    if (not v_orig or "skaps" in v_orig.lower()) and valid_vendor:
        logger.debug(f"  [Vendor Inherit] Replacing '{v_orig}' with inherited vendor '{valid_vendor}'")
        header["Vendor Name"] = valid_vendor

    lines  = block_text.split("\n")

    header_idx = find_item_table_header(lines)
    raw_items = []

    if header_idx >= 0:
        raw_items = extract_items_from_table(lines, header_idx)
        logger.debug(f"  Table found at line {header_idx}; raw items: {len(raw_items)}")

    if not raw_items:
        logger.debug("  No item table detected -> Running fallback line-by-line item extraction...")
        raw_items = extract_items_fallback(lines)
        logger.debug(f"  Fallback raw items extracted: {len(raw_items)}")

    # Resolve HSNs from the block if any items are missing them
    if raw_items:
        missing_hsn = any(not it.get("hsn") for it in raw_items)
        if missing_hsn:
            hsns = extract_all_hsns(block_text)
            if hsns:
                if len(hsns) == 1:
                    for it in raw_items:
                        if not it.get("hsn"):
                            it["hsn"] = hsns[0]
                elif len([it for it in raw_items if not it.get("hsn")]) == len(hsns):
                    hsns_idx = 0
                    for it in raw_items:
                        if not it.get("hsn") and hsns_idx < len(hsns):
                            it["hsn"] = hsns[hsns_idx]
                            hsns_idx += 1

    if not raw_items:
        logger.info("  No valid item rows found → creating Partial record.")
        return [
            make_record(
                source_file, header,
                status="Partial",
                remarks=f"No item table detected; {method} used",
            )
        ]

    records: list[dict] = []
    for item_dict in raw_items:
        remark = f"Extracted via {method}"

        records.append(
            make_record(
                source_file, header,
                item=item_dict["item"],
                hsn=item_dict.get("hsn", ""),
                qty=item_dict["qty"],
                price=item_dict["price"],
                status="Success",
                remarks=remark,
            )
        )

    return records


# ─────────────────────────────────────────────────────────────────────────────
# STEP 9: EXTRACT RECORDS FROM A FULL PDF TEXT
# ─────────────────────────────────────────────────────────────────────────────

def extract_records(text: str, text_fallback: str, source_file: str, method: str) -> list[dict]:
    """
    Split the full PDF text into invoice blocks (handles multi-invoice PDFs),
    then run the generic extractor on each block.
    """
    blocks_indexed = split_into_invoice_blocks_with_pages(text)
    logger.info(f"  Invoice blocks detected: {len(blocks_indexed)}")

    pages_fallback = PAGE_SEP.split(text_fallback)
    pages_fallback = [p.strip() for p in pages_fallback]

    # Find the first valid vendor name across all blocks (original and fallback)
    valid_vendor = ""
    for block_pages in blocks_indexed:
        block_text = "\n".join([p for _, p in block_pages])
        h = extract_header_generic(block_text)
        v = h.get("Vendor Name", "").strip()
        if v and "skaps" not in v.lower():
            valid_vendor = v
            break
    if not valid_vendor:
        for p_fallback in pages_fallback:
            h_fallback = extract_header_generic(p_fallback)
            v_fallback = h_fallback.get("Vendor Name", "").strip()
            if v_fallback and "skaps" not in v_fallback.lower():
                valid_vendor = v_fallback
                break

    all_records: list[dict] = []
    for i, block_pages in enumerate(blocks_indexed):
        block_text = "\n".join([p for _, p in block_pages])
        
        fallback_page_texts = []
        for idx, _ in block_pages:
            if idx < len(pages_fallback):
                fallback_page_texts.append(pages_fallback[idx])
        block_fallback = "\n".join(fallback_page_texts)

        logger.debug(f"  Processing block {i + 1}/{len(blocks_indexed)} …")
        records = extract_generic_items(block_text, block_fallback, source_file, method, valid_vendor)
        all_records.extend(records)

    return all_records


# ─────────────────────────────────────────────────────────────────────────────
# STEP 10: PROCESS A SINGLE PDF FILE
# ─────────────────────────────────────────────────────────────────────────────

def process_pdf(pdf_path: Path) -> list[dict]:
    """
    Full pipeline for one PDF:
      1. Extract text (pdfplumber or OCR)
      2. Clean and save raw text
      3. Split into invoice blocks and extract records
    """
    logger.info(f"─── Processing: {pdf_path.name}")
    text_300, text_200, method = get_pdf_text(pdf_path)
    
    # Save the original raw extracted text (with HTML tables if Paddle-OCR)
    save_extracted_text(pdf_path.name, text_300, method)

    # Preprocess Paddle-OCR HTML tables into plain text line layout for regex compatibility
    if method == "paddleocr":
        text_300 = convert_html_tables_to_text(text_300)
        text_200 = convert_html_tables_to_text(text_200)

    text_300 = clean_text(text_300)
    text_200 = clean_text(text_200)

    records = extract_records(text_300, text_200, pdf_path.name, method)
    logger.info(f"  → {len(records)} record(s) from {pdf_path.name}")
    return records


# ─────────────────────────────────────────────────────────────────────────────
# STEP 11: EXCEL WRITING AND FORMATTING
# ─────────────────────────────────────────────────────────────────────────────

def load_existing_records(excel_path: Path) -> list[dict]:
    """Load existing records from Excel if the file exists."""
    if not excel_path.exists():
        return []
    try:
        df = pd.read_excel(str(excel_path), sheet_name="Invoice Records")
        df = df.fillna("")
        return df.to_dict(orient="records")
    except Exception as e:
        logger.error(f"Failed to load existing records: {e}")
        return []


def write_excel(records: list[dict], output_path: Path) -> None:
    """Write all records to a formatted Excel file (sheet: Invoice Records)."""
    df = pd.DataFrame(records, columns=EXCEL_COLUMNS)
    df.to_excel(str(output_path), index=False,
                sheet_name="Invoice Records", engine="openpyxl")

    wb = load_workbook(str(output_path))
    ws = wb["Invoice Records"]

    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_font  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    for col_idx, col_name in enumerate(EXCEL_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len    = len(col_name)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

    data_align = Alignment(vertical="top", wrap_text=True)
    alt_fill   = PatternFill("solid", fgColor="EEF2F7")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_align
            if cell.row % 2 == 0:
                cell.fill = alt_fill

    wb.save(str(output_path))
    logger.info(f"Excel saved → {output_path}")


def save_text_file(records: list[dict], output_path: Path) -> None:
    """Saves a structured TXT summary of the extracted records (same structure as Excel)."""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("==================================================\n")
        f.write("            EXTRACTED INVOICE RECORDS             \n")
        f.write("==================================================\n\n")

        for idx, rec in enumerate(records, start=1):
            f.write(f"Record #{idx}:\n")
            f.write(f"  Source File:    {rec.get('Source File', '')}\n")
            f.write(f"  Vendor Name:    {rec.get('Vendor Name', '')}\n")
            f.write(f"  Invoice Number: {rec.get('Invoice Number', '')}\n")
            f.write(f"  Invoice Date:   {rec.get('Invoice Date', '')}\n")
            f.write(f"  PO Number:      {rec.get('PO Number', '')}\n")
            f.write(f"  Item:           {rec.get('Item', '')}\n")
            f.write(f"  HSN Number:     {rec.get('HSN Number', '')}\n")
            f.write(f"  Quantity:       {rec.get('Quantity', '')}\n")
            f.write(f"  Basic Price:    {rec.get('Basic Price excluding GST', '')}\n")
            f.write(f"  Status:         {rec.get('Extraction Status', '')}\n")
            f.write(f"  Remarks:        {rec.get('Remarks', '')}\n")
            f.write("-" * 50 + "\n")


# ─────────────────────────────────────────────────────────────────────────────
# STEP 12: MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    for folder in [INVOICES_DIR, EXTRACTED_TEXT_DIR, OUTPUT_DIR,
                   PROCESSED_DIR, ERROR_DIR]:
        folder.mkdir(parents=True, exist_ok=True)

    pdf_files = sorted(INVOICES_DIR.glob("*.pdf"))
    if not pdf_files:
        logger.warning(f"No PDF files found in '{INVOICES_DIR}'. Exiting.")
        return

    logger.info(f"Found {len(pdf_files)} PDF file(s) to process.")

    all_records: list[dict] = []
    processed_count = 0
    failed_pdfs: list[str] = []

    for pdf_path in pdf_files:
        try:
            records = process_pdf(pdf_path)
            all_records.extend(records)
            processed_count += 1
            shutil.move(str(pdf_path), str(PROCESSED_DIR / pdf_path.name))
            logger.info(f"  '{pdf_path.name}' → processed/")
        except Exception as exc:
            logger.error(f"  FAILED '{pdf_path.name}': {exc}", exc_info=True)
            failed_pdfs.append(pdf_path.name)
            all_records.append(
                make_record(pdf_path.name, {}, status="Failed",
                            remarks=f"ERROR: {exc}")
            )
            try:
                shutil.move(str(pdf_path), str(ERROR_DIR / pdf_path.name))
            except Exception:
                pass

    if all_records:
        existing_records = load_existing_records(OUTPUT_EXCEL)
        new_files = {r["Source File"] for r in all_records}
        combined_records = [r for r in existing_records if r.get("Source File") not in new_files]
        combined_records.extend(all_records)
        
        write_excel(combined_records, OUTPUT_EXCEL)
        save_text_file(combined_records, OUTPUT_TXT)
    else:
        logger.warning("No records extracted – Excel/TXT not created.")

    success = [r for r in all_records if r["Extraction Status"] == "Success"]
    partial = [r for r in all_records if r["Extraction Status"] == "Partial"]
    failed  = [r for r in all_records if r["Extraction Status"] == "Failed"]

    log_and_print("\n" + "=" * 58)
    log_and_print("         INVOICE EXTRACTION SUMMARY")
    log_and_print("=" * 58)
    log_and_print(f"  PDFs found          : {len(pdf_files)}")
    log_and_print(f"  PDFs processed OK   : {processed_count}")
    log_and_print(f"  PDFs failed         : {len(failed_pdfs)}")
    log_and_print(f"  Total rows          : {len(all_records)}")
    log_and_print(f"    Success           : {len(success)}")
    log_and_print(f"    Partial           : {len(partial)}")
    log_and_print(f"    Failed            : {len(failed)}")
    if failed_pdfs:
        log_and_print("\n  Failed files:")
        for f in failed_pdfs:
            log_and_print(f"    - {f}")
    log_and_print(f"\n  Output Excel  : {OUTPUT_EXCEL}")
    log_and_print(f"  Extracted text: {EXTRACTED_TEXT_DIR}")
    log_and_print(f"  Log file      : {LOG_FILE}")
    log_and_print("=" * 58 + "\n")


if __name__ == "__main__":
    main()
